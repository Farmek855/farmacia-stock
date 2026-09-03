# -*- coding: utf-8 -*-
"""
Exporta los productos con stock disponible desde ROTACIONES_E_INVENTARIO.xlsx/.xlsm
hacia docs/data.json, y sube el cambio a GitHub automaticamente (add + commit + push).

REQUISITOS (una sola vez):
    pip install openpyxl
    Tener Git para Windows instalado y el repositorio ya clonado localmente
    (ver guia de configuracion de GitHub que te paso aparte).

USO:
    python exportar_stock.py

Puedes ejecutarlo a mano con doble clic (usando el .bat de ejemplo), desde una
tarea programada de Windows, o encadenarlo al final de tu macro de Excel con
un Shell() que llame a este script.
"""

import json
import subprocess
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

# ============================================================================
# CONFIGURACION - AJUSTA ESTAS 3 RUTAS SEGUN TU COMPUTADOR
# ============================================================================

# Ruta al Excel real (el que sincroniza Google Drive de escritorio).
RUTA_EXCEL = r"G:\Mi unidad\WEBPAGE\ROTACIONES E INVENTARIO.xlsm"

# Carpeta LOCAL donde clonaste el repositorio de GitHub (fuera de Google Drive,
# para que git y Google Drive no se pisen entre si).
RUTA_REPO = r"C:\Users\mekni\OneDrive\Escritorio\webpage\farmacia-stock"

# ============================================================================
# NO ES NECESARIO EDITAR NADA MAS ABAJO DE ESTA LINEA
# ============================================================================

COL_CODIGO = "B"
COL_TIPO = "C"
COL_CATEGORIA = "D"
COL_NOMBRE = "F"          # Marca / nombre comercial del producto
COL_PRESENTACION = "G"
COL_DESCRIPCION = "H"     # Descripcion generica (principio activo)
COL_SALDO = "J"

FILA_INICIO = 5


def leer_productos_con_stock(ruta_excel: str):
    wb = openpyxl.load_workbook(ruta_excel, data_only=True)
    ws = wb["Hoja1"]

    productos = []
    fila = FILA_INICIO
    while True:
        codigo = ws[f"{COL_CODIGO}{fila}"].value
        if codigo is None or str(codigo).strip() == "":
            # Se asume que dos filas vacias seguidas marcan el fin de los datos
            siguiente = ws[f"{COL_CODIGO}{fila + 1}"].value
            if siguiente is None or str(siguiente).strip() == "":
                break
            fila += 1
            continue

        saldo = ws[f"{COL_SALDO}{fila}"].value
        try:
            saldo = float(saldo) if saldo is not None else 0
        except (TypeError, ValueError):
            saldo = 0

        if saldo > 0:
            productos.append({
                "codigo": str(codigo).strip(),
                "nombre": (ws[f"{COL_NOMBRE}{fila}"].value or "").strip()
                    if isinstance(ws[f"{COL_NOMBRE}{fila}"].value, str)
                    else (ws[f"{COL_NOMBRE}{fila}"].value or ""),
                "descripcion": (ws[f"{COL_DESCRIPCION}{fila}"].value or "").strip()
                    if isinstance(ws[f"{COL_DESCRIPCION}{fila}"].value, str)
                    else (ws[f"{COL_DESCRIPCION}{fila}"].value or ""),
                "saldo": int(saldo) if saldo == int(saldo) else saldo,
                "tipo": (ws[f"{COL_TIPO}{fila}"].value or "").strip()
                    if isinstance(ws[f"{COL_TIPO}{fila}"].value, str)
                    else (ws[f"{COL_TIPO}{fila}"].value or ""),
                "categoria": (ws[f"{COL_CATEGORIA}{fila}"].value or "").strip()
                    if isinstance(ws[f"{COL_CATEGORIA}{fila}"].value, str)
                    else (ws[f"{COL_CATEGORIA}{fila}"].value or ""),
                "presentacion": (ws[f"{COL_PRESENTACION}{fila}"].value or "").strip()
                    if isinstance(ws[f"{COL_PRESENTACION}{fila}"].value, str)
                    else (ws[f"{COL_PRESENTACION}{fila}"].value or ""),
            })
        fila += 1

    wb.close()
    return productos


def escribir_data_json(productos, ruta_repo: str):
    ruta_docs = Path(ruta_repo) / "docs"
    ruta_docs.mkdir(parents=True, exist_ok=True)
    ruta_json = ruta_docs / "data.json"

    payload = {
        "generado": datetime.now().strftime("%d-%m-%Y %H:%M"),
        "productos": productos,
    }

    with open(ruta_json, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    return ruta_json


def subir_a_github(ruta_repo: str):
    def run(cmd):
        return subprocess.run(
            cmd, cwd=ruta_repo, capture_output=True, text=True, shell=False
        )

    run(["git", "add", "docs/data.json"])

    mensaje = f"Actualizacion de stock - {datetime.now().strftime('%d-%m-%Y %H:%M')}"
    resultado_commit = run(["git", "commit", "-m", mensaje])

    if "nothing to commit" in (resultado_commit.stdout + resultado_commit.stderr).lower():
        print("Sin cambios en el stock desde la ultima actualizacion. No se hizo push.")
        return

    resultado_push = run(["git", "push"])
    if resultado_push.returncode != 0:
        print("ERROR al hacer push a GitHub:")
        print(resultado_push.stderr)
        sys.exit(1)

    print("Cambios subidos a GitHub correctamente.")


def main():
    if not Path(RUTA_EXCEL).exists():
        print(f"ERROR: no se encontro el Excel en: {RUTA_EXCEL}")
        print("Edita la variable RUTA_EXCEL al inicio de este script.")
        sys.exit(1)

    if not (Path(RUTA_REPO) / ".git").exists():
        print(f"ERROR: {RUTA_REPO} no parece ser un repositorio git valido.")
        print("Edita la variable RUTA_REPO al inicio de este script.")
        sys.exit(1)

    print("Leyendo productos con stock desde el Excel...")
    productos = leer_productos_con_stock(RUTA_EXCEL)
    print(f"Se encontraron {len(productos)} productos con stock.")

    ruta_json = escribir_data_json(productos, RUTA_REPO)
    print(f"Archivo generado: {ruta_json}")

    print("Subiendo cambios a GitHub...")
    subir_a_github(RUTA_REPO)


if __name__ == "__main__":
    main()
