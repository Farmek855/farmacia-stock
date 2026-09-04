# -*- coding: utf-8 -*-
"""
Version en la nube de exportar_stock.py, pensada para correr dentro de
GitHub Actions (no en tu computador). Descarga el Excel directamente desde
Google Drive usando el ID del archivo, sincroniza fotos y documentos desde
una carpeta de Drive, y genera docs/data.json + docs/images + docs/docs.

Variables de entorno requeridas:
  DRIVE_FILE_ID        - ID del archivo Excel (Settings > Variables)
  DRIVE_MEDIA_FOLDER_ID - ID de la carpeta de Drive con fotos/documentos (Variables)
  DRIVE_API_KEY        - API key de Google Drive API (Settings > Secrets)
"""

import json
import os
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
import requests

COL_CODIGO = "B"
COL_TIPO = "C"
COL_CATEGORIA = "D"
COL_NOMBRE = "F"
COL_PRESENTACION = "G"
COL_DESCRIPCION = "H"
COL_SALDO = "J"
COL_FECHA_INGRESO = "Z"

FILA_INICIO = 5
DIAS_CONSIDERADO_NUEVO = 28

RUTA_TEMPORAL = "planilla_descargada.xlsm"
RUTA_JSON = Path("docs/data.json")
CARPETA_IMAGENES = Path("docs/images")
CARPETA_DOCS = Path("docs/docs")

EXTENSIONES_IMAGEN = {".jpg", ".jpeg", ".png", ".webp"}
EXTENSIONES_DOC = {".pdf", ".txt", ".html", ".htm"}


def descargar_de_google_drive(file_id: str, destino: str):
    """
    Descarga un archivo publico ('cualquiera con el enlace') desde Google Drive,
    manejando la pantalla de confirmacion que Google muestra para archivos
    con macros o de mayor tamano.
    """
    url = "https://docs.google.com/uc?export=download"
    session = requests.Session()

    respuesta = session.get(url, params={"id": file_id}, stream=True)

    token = None
    for clave, valor in respuesta.cookies.items():
        if clave.startswith("download_warning"):
            token = valor
            break

    if token:
        respuesta = session.get(
            url, params={"id": file_id, "confirm": token}, stream=True
        )

    if respuesta.status_code != 200:
        print(f"ERROR: Google Drive respondio con codigo {respuesta.status_code}")
        sys.exit(1)

    with open(destino, "wb") as f:
        for chunk in respuesta.iter_content(32768):
            if chunk:
                f.write(chunk)


def valor_texto(celda):
    v = celda.value
    if isinstance(v, str):
        return v.strip()
    return v or ""


def leer_productos_con_stock(ruta_excel: str):
    wb = openpyxl.load_workbook(ruta_excel, data_only=True)
    ws = wb["Hoja1"]

    productos = []
    fila = FILA_INICIO
    while True:
        codigo = ws[f"{COL_CODIGO}{fila}"].value
        if codigo is None or str(codigo).strip() == "":
            siguiente = ws[f"{COL_CODIGO}{fila + 1}"].value
            if siguiente is None or str(siguiente).strip() == "":
                break
            fila += 1
            continue

        saldo = ws[f"{COL_SALDO}{fila}"].value
        try:
            saldo = float(saldo) if saldo is not None else 0
        except (TypeError, ValueError):
            saldo = 0

        if saldo > 0:
            fecha_ingreso_val = ws[f"{COL_FECHA_INGRESO}{fila}"].value
            es_nuevo = False
            if isinstance(fecha_ingreso_val, datetime):
                dias = (datetime.now() - fecha_ingreso_val).days
                es_nuevo = 0 <= dias <= DIAS_CONSIDERADO_NUEVO

            productos.append({
                "codigo": str(codigo).strip(),
                "nombre": valor_texto(ws[f"{COL_NOMBRE}{fila}"]),
                "descripcion": valor_texto(ws[f"{COL_DESCRIPCION}{fila}"]),
                "saldo": int(saldo) if saldo == int(saldo) else saldo,
                "nuevo": es_nuevo,
                "tipo": valor_texto(ws[f"{COL_TIPO}{fila}"]),
                "categoria": valor_texto(ws[f"{COL_CATEGORIA}{fila}"]),
                "presentacion": valor_texto(ws[f"{COL_PRESENTACION}{fila}"]),
            })
        fila += 1

    wb.close()
    return productos


def listar_archivos_carpeta_drive(folder_id: str, api_key: str):
    """
    Lista los archivos dentro de una carpeta publica de Google Drive,
    usando la API de Drive con una API key (sin OAuth).
    """
    archivos = []
    url = "https://www.googleapis.com/drive/v3/files"
    page_token = None

    while True:
        params = {
            "q": f"'{folder_id}' in parents and trashed = false",
            "key": api_key,
            "fields": "nextPageToken, files(id, name)",
            "pageSize": 1000,
        }
        if page_token:
            params["pageToken"] = page_token

        r = requests.get(url, params=params)
        if r.status_code != 200:
            print(f"ERROR listando carpeta de Drive: {r.status_code} - {r.text}")
            sys.exit(1)

        data = r.json()
        archivos.extend(data.get("files", []))
        page_token = data.get("nextPageToken")
        if not page_token:
            break

    return archivos


def descargar_archivo_drive_api(file_id: str, api_key: str, destino: Path):
    url = f"https://www.googleapis.com/drive/v3/files/{file_id}"
    r = requests.get(url, params={"alt": "media", "key": api_key}, stream=True)
    if r.status_code != 200:
        print(f"  ADVERTENCIA: no se pudo descargar {destino.name} ({r.status_code})")
        return False
    with open(destino, "wb") as f:
        for chunk in r.iter_content(32768):
            if chunk:
                f.write(chunk)
    return True


def extraer_codigo_y_tipo(nombre_archivo: str):
    """
    'CODIGOimg.jpg' -> (CODIGO, 'img')
    'CODIGOdoc.pdf' -> (CODIGO, 'doc')
    Devuelve (None, None) si no calza con el patron esperado.
    """
    base, ext = os.path.splitext(nombre_archivo)
    ext = ext.lower()

    if base.endswith("img") and ext in EXTENSIONES_IMAGEN:
        return base[:-3], "img", ext
    if base.endswith("doc") and ext in EXTENSIONES_DOC:
        return base[:-3], "doc", ext

    return None, None, None


def sincronizar_media(folder_id: str, api_key: str):
    """
    Descarga fotos y documentos desde la carpeta de Drive.
    Devuelve dos diccionarios: {codigo: ruta_relativa_imagen}, {codigo: ruta_relativa_doc}
    """
    CARPETA_IMAGENES.mkdir(parents=True, exist_ok=True)
    CARPETA_DOCS.mkdir(parents=True, exist_ok=True)

    # Git no rastrea carpetas vacias: dejamos un placeholder para que
    # 'git add' nunca falle aunque todavia no haya fotos/documentos subidos.
    (CARPETA_IMAGENES / ".gitkeep").touch(exist_ok=True)
    (CARPETA_DOCS / ".gitkeep").touch(exist_ok=True)

    imagenes = {}
    documentos = {}

    archivos = listar_archivos_carpeta_drive(folder_id, api_key)
    print(f"Archivos encontrados en la carpeta de Drive: {len(archivos)}")

    for archivo in archivos:
        codigo, tipo, ext = extraer_codigo_y_tipo(archivo["name"])
        if codigo is None:
            print(f"  Omitido (no calza con el patron CODIGOimg/CODIGOdoc): {archivo['name']}")
            continue

        if tipo == "img":
            destino = CARPETA_IMAGENES / f"{codigo}{ext}"
            if descargar_archivo_drive_api(archivo["id"], api_key, destino):
                imagenes[codigo] = f"images/{codigo}{ext}"
        elif tipo == "doc":
            destino = CARPETA_DOCS / f"{codigo}{ext}"
            if descargar_archivo_drive_api(archivo["id"], api_key, destino):
                documentos[codigo] = f"docs/{codigo}{ext}"

    return imagenes, documentos


def main():
    file_id = os.environ.get("DRIVE_FILE_ID")
    if not file_id:
        print("ERROR: falta la variable de entorno DRIVE_FILE_ID")
        sys.exit(1)

    print("Descargando planilla desde Google Drive...")
    descargar_de_google_drive(file_id, RUTA_TEMPORAL)

    print("Leyendo productos con stock...")
    productos = leer_productos_con_stock(RUTA_TEMPORAL)
    print(f"Se encontraron {len(productos)} productos con stock.")

    media_folder_id = os.environ.get("DRIVE_MEDIA_FOLDER_ID")
    api_key = os.environ.get("DRIVE_API_KEY")
    imagenes, documentos = {}, {}
    CARPETA_IMAGENES.mkdir(parents=True, exist_ok=True)
    CARPETA_DOCS.mkdir(parents=True, exist_ok=True)
    (CARPETA_IMAGENES / ".gitkeep").touch(exist_ok=True)
    (CARPETA_DOCS / ".gitkeep").touch(exist_ok=True)
    if media_folder_id and api_key:
        print("Sincronizando fotos y documentos desde Drive...")
        imagenes, documentos = sincronizar_media(media_folder_id, api_key)
        print(f"Fotos sincronizadas: {len(imagenes)} | Documentos sincronizados: {len(documentos)}")
    else:
        print("Sin DRIVE_MEDIA_FOLDER_ID / DRIVE_API_KEY: se omite sincronizacion de fotos y documentos.")

    for p in productos:
        cod = p["codigo"]
        if cod in imagenes:
            p["imagen"] = imagenes[cod]
        if cod in documentos:
            p["documento"] = documentos[cod]

    RUTA_JSON.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "generado": datetime.now().strftime("%d-%m-%Y %H:%M"),
        "productos": productos,
    }
    with open(RUTA_JSON, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    print(f"Archivo generado: {RUTA_JSON}")
    os.remove(RUTA_TEMPORAL)


if __name__ == "__main__":
    main()
